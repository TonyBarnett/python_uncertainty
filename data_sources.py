from itertools import product
import os
from openpyxl import load_workbook
import pymongo
import pymssql
import regex

from string import ascii_uppercase as uppercase

DROPBOX_ROOT = os.getenv("dropboxRoot") + "\\"
UNCERTAINTY_DIR = DROPBOX_ROOT + "IO Model source data\\Source data\\Uncertainty_files\\"
SOURCE_FILE_DIR = DROPBOX_ROOT + "IO Model source data\\Source data\\Source data files\\"


def _is_number(value: str) -> bool:
    if not value:
        return False
    # Let's assume that the value has potentially been cleaned for mongo, i.e. .s have become _s
    val = value.replace('_', '.')
    try:
        float(val)
    except ValueError:
        return False
    return True


def _read_from_sql(query: str, params: tuple=None, db: str=None, server='localhost', uid='sa', pw='deter101!'):
    with pymssql.connect(server, uid, pw, db) as conn:
        with conn.cursor() as cursor:
            # The star says you want a list of parameters for the format.
            cursor.execute(query if not params else query.format(*params))
            return cursor.fetchall()


def _base_26_number_plus_1(number: list) -> list:
    number.insert(0, 0)
    index = len(number)
    for n in range(index - 1, -1, -1):
        number[n] += 1
        # keep incrementing elements until we find one that's less than 26
        if number[n] <= 26:
            if number[0] == 0:
                number.pop(0)
            return number
        number[n] = 1


def _convert_str_to_base_26(value: str) -> list:
    # ord("A")==65
    return [ord(c) - 64 for c in value]


def _check_end_after_start(start, end) -> None:
    reverse_start = start[::-1]
    reverse_end = end[::-1]
    e = sum((26 ** i) * ord(c) for i, c in enumerate(reverse_end))
    s = sum((26 ** i) * ord(c) for i, c in enumerate(reverse_start))
    if s > e:
        raise ValueError("start value {0} before end value {1}".format(start, end))

def _check_start_end_acceptable(start: str, end: str) -> None:
    """
    check that the start and end strings for the char getter are acceptable.
    TODO: rename this to something more sensible
    """

    char_regex = regex.compile("[A-Z]+")

    if not char_regex.fullmatch(start) or not char_regex.fullmatch(end):
        raise ValueError("start and end must be characters")

    _check_end_after_start(start, end)

def _get_chars_in_range(start: str, end: str):
    """
Returns all characters between start, and end in an excel type manner (so Y Z AA AB AC...). Assumes start is before end
    @param start:
    @param end:
    @return list of strings:
    """
    if not start or not end:
        return []

    end = end.upper()
    start = start.upper()

    _check_start_end_acceptable(start, end)

    range_builder = [start]

    start_list = _convert_str_to_base_26(start)
    end_list = _convert_str_to_base_26(end)

    # we always want to add start_list value, we're fairly sure that end >= start
    while start_list != end_list:
        start_list = _base_26_number_plus_1(start_list)
        # convert this funny base 26 list to a string
        range_builder.append("".join([chr(x + 64) for x in start_list]))
    return range_builder


def _get_cell_in_range(start, end) -> iter:
    cell_regex = regex.compile("([A-Z]+)([0-9]+)")
    start_match = cell_regex.match(start)
    end_match = cell_regex.match(end)
    start_char = start_match.group(1)
    start_num = int(start_match.group(2))
    end_char = end_match.group(1)
    end_num = int(end_match.group(2))

    chars = _get_chars_in_range(start_char, end_char)
    nums = range(start_num, end_num + 1)

    return (c + str(n) for c in chars for n in nums)


def _get_data_from_workbook(workbook, worksheet_name: str, start_cell: str, end_cell: str) -> tuple:
    ws = workbook.get_sheet_by_name(worksheet_name)
    cell_range = _get_cell_in_range(start_cell, end_cell)
    values = list()
    for cell in cell_range:
        values.append(str(ws[cell].value))
    return tuple(values)


def _clean_uk_supply_error_value(value):
    if value == "*":
        return 0

    if value == "-":
        return "c"

    return value


def _build_query(system_id: str, value_id: str, used: bool) -> str:
    where = list()
    select = "v.strSystemId, v.strValue, v.strDescription "
    from_ = "clasValue v"

    inner_join = ""
    if used:
        inner_join = " clasSystem s ON s.strName = v.strSystemId"
        where.append("s.bolUsed = {0}".format(1 if used else 0))

    if system_id:
        if not value_id:
            raise ValueError("cannot have a value_id without a system_id")
        where.append(" v.strSystemId = %s")

        if value_id:
            where.append(" v.strValue = %s")
    query = "SELECT {0} FROM {1} ".format(select, from_)

    if inner_join:
        query += "INNER JOIN {0} ".format(inner_join)
    # we have to have four letters so we have something to strip off at the end if there are no conditions
    # feels a little hacky but this isn't a big thing
    where_string = "blah"
    if used or system_id:
        where_string = "WHERE "
        for w in where:
            where_string += " {0} AND".format(w)

    # trim the last AND
    return query + where_string[:-4]


def get_map(map_collection: str='Plain_KNN_Without_Ancestors_k_3'):
    collection = pymongo.MongoClient('10.10.20.37').ClassificationMap[map_collection]
    mapping = dict()
    for map_ in collection.find():
        system = map_["_id"]["system"]
        value = map_["_id"]["value"]
        mapped_values = map_["map"]
        if system not in mapping:
            mapping[system] = dict()
        mapping[system][value] = mapped_values
    return mapping


def get_classification_systems_data(system_id: str=None, value_id: str=None, used: bool=None):
    """ I'm sorry, this is really badly made but I can't be bothered to make it any better
    :param system_id:
    :param value_id:
    :param used:
    :return:
    """
    query = _build_query(system_id, value_id, used)
    class_sys_data = dict()
    for sys_id, val_id, description in _read_from_sql(query, tuple([system_id, value_id]), db="ClassificationSystems"):
        if sys_id not in class_sys_data:
            class_sys_data[sys_id] = dict()

        class_sys_data[sys_id][val_id] = description
    return class_sys_data


def get_uk_supply(year) -> dict:
    totals_file = SOURCE_FILE_DIR + "UKSupply_source.xlsx"
    wb = load_workbook(totals_file, read_only=True)

    keys = _get_data_from_workbook(wb, "sup{0}".format(str(year)[-2:]), "D6", "BN6")
    data = _get_data_from_workbook(wb, "sup{0}".format(str(year)[-2:]), "D73", "BN73")
    values = dict()

    for i, key in enumerate(keys):
        values[key] = data[i]

    return values


def get_uk_supply_error(year) -> dict:
    file_name = UNCERTAINTY_DIR + "Annual Business Survey_2008-2013.xlsx"

    wb = load_workbook(file_name)
    labels = _get_data_from_workbook(wb, "Quality Measures", "B14", "B5532")
    years = _get_data_from_workbook(wb, "Quality Measures", "D14", "D5532")
    raw_data = _get_data_from_workbook(wb, "Quality Measures", "G14", "G5532")

    data = dict()
    label = ""
    for i, item in enumerate(raw_data):
        if years[i] != str(year):
            continue
        if labels[i]:
            label = labels[i]
        data[label] = _clean_uk_supply_error_value(raw_data[i])

    return data


def get_eu_emissions_error_from_file(file_name: str,
                                     data_start_cell: str,
                                     data_end_cell: str,
                                     error_start_cell: str,
                                     error_end_cell: str) -> list:
    wb = load_workbook(file_name)

    raw_data = _get_data_from_workbook(wb, "Data", data_start_cell, data_end_cell)
    error_data = _get_data_from_workbook(wb, "Data", error_start_cell, error_end_cell)
    data = list()

    for i, item in enumerate(raw_data):
        if _is_number(error_data[i]):
            data.append((item, error_data[i]))
    return data


def get_eu_emissions_error() -> tuple:
    data = list()
    file_name = UNCERTAINTY_DIR + "UK_emissions_source_2004.xlsx"
    data += get_eu_emissions_error_from_file(file_name, "B2", "B73", "E2", "E73")

    file_name = UNCERTAINTY_DIR + "UK_emissions_source_2010.xlsx"
    data += get_eu_emissions_error_from_file(file_name, "D2", "D63", "G2", "G63")

    file_name = UNCERTAINTY_DIR + "UK_emissions_source_2011.xlsx"
    data += get_eu_emissions_error_from_file(file_name, "D2", "D62", "G2", "G62")

    file_name = UNCERTAINTY_DIR + "UK_emissions_source_2012.xlsx"
    data += get_eu_emissions_error_from_file(file_name, "D2", "D70", "G2", "G70")

    return tuple(data)