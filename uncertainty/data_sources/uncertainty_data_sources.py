import os
from openpyxl import load_workbook
from .excel import _get_data_from_workbook

DROPBOX_ROOT = os.getenv("dropboxRoot") + "\\"
UNCERTAINTY_DIR = DROPBOX_ROOT + "IO Model source data\\Source data\\Uncertainty_files\\"
SOURCE_FILE_DIR = DROPBOX_ROOT + "IO Model source data\\Source data\\Source data files\\"


def _is_number(value: str) -> bool:
    if not value:
        return False
    # Let's assume that the value has potentially been cleaned for mongo, i.e. .s have become _s
    val = value.replace('_', '.')
    try:
        float(val)
    except ValueError:
        return False
    return True


def _clean_uk_supply_error_value(value):
    if value == "*":
        return 0

    if value == "-":
        return "c"

    return value


def get_uk_supply(year) -> dict:
    totals_file = SOURCE_FILE_DIR + "UKSupply_source.xlsx"
    wb = load_workbook(totals_file, read_only=True)

    keys = _get_data_from_workbook(wb, "sup{0}".format(str(year)[-2:]), "D6", "BN6")
    data = _get_data_from_workbook(wb, "sup{0}".format(str(year)[-2:]), "D73", "BN73")
    values = dict()

    for i, key in enumerate(keys):
        values[key] = data[i]

    return values


def get_uk_supply_error(year) -> dict:
    file_name = UNCERTAINTY_DIR + "Annual Business Survey_2008-2013.xlsx"

    wb = load_workbook(file_name)
    labels = _get_data_from_workbook(wb, "Quality Measures", "B14", "B5532")
    years = _get_data_from_workbook(wb, "Quality Measures", "D14", "D5532")
    raw_data = _get_data_from_workbook(wb, "Quality Measures", "G14", "G5532")

    data = dict()
    label = ""
    for i, item in enumerate(raw_data):
        if years[i] != str(year):
            continue
        if labels[i]:
            label = labels[i]
        data[label] = _clean_uk_supply_error_value(raw_data[i])

    return data


def get_eu_emissions_error_from_file(file_name: str,
                                     data_start_cell: str,
                                     data_end_cell: str,
                                     error_start_cell: str,
                                     error_end_cell: str) -> list:
    wb = load_workbook(file_name)

    raw_data = _get_data_from_workbook(wb, "Data", data_start_cell, data_end_cell)
    error_data = _get_data_from_workbook(wb, "Data", error_start_cell, error_end_cell)
    data = list()

    for i, item in enumerate(raw_data):
        if _is_number(error_data[i]):
            data.append((item, error_data[i]))
    return data


def get_eu_emissions_error() -> tuple:
    data = list()
    file_name = UNCERTAINTY_DIR + "UK_emissions_source_2004.xlsx"
    data += get_eu_emissions_error_from_file(file_name, "B2", "B73", "E2", "E73")

    file_name = UNCERTAINTY_DIR + "UK_emissions_source_2010.xlsx"
    data += get_eu_emissions_error_from_file(file_name, "D2", "D63", "G2", "G63")

    file_name = UNCERTAINTY_DIR + "UK_emissions_source_2011.xlsx"
    data += get_eu_emissions_error_from_file(file_name, "D2", "D62", "G2", "G62")

    file_name = UNCERTAINTY_DIR + "UK_emissions_source_2012.xlsx"
    data += get_eu_emissions_error_from_file(file_name, "D2", "D70", "G2", "G70")

    return tuple(data)
